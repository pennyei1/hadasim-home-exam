import csv
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import load_workbook
import os

"""
*****************************************************
section A: counting occurrences of Errors form logs.txt
*****************************************************
"""

LINES_PER_CHUNK = 50000


# 1. code for splitting the file into smaller parts
def split_lines(file):
    """
        the function split the file into smaller files - makes it efficient in memory. in each time only
        one file is in memory, and not the whole file
        :param file: the original file
    """

    chunk_index = 0
    cur_chunk = []

    workbook = load_workbook(filename=file, read_only=True)
    sheet = workbook.active

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        line = ' '.join(str(cell) if cell is not None else '' for cell in row) + '\n'
        cur_chunk.append(line)

        if row_index % LINES_PER_CHUNK == 0:
            with open(f'chunk_{chunk_index}.txt', 'w', encoding='utf-8') as chunk:
                chunk.writelines(cur_chunk)
            cur_chunk = []
            chunk_index += 1

    # Write remaining lines (if any)
    if cur_chunk:
        with open(f'chunk_{chunk_index}.txt', 'w', encoding='utf-8') as chunk:
            chunk.writelines(cur_chunk)
        chunk_index += 1

    return chunk_index


# 2. counts the frequency for each error
def count_occurrences_in_file(error_dictionary, num_file):
    """
        the function counts the number of occurrences for each error
        :param error_dictionary:
        :param num_file: the current file

    """
    with open(f'chunk_{num_file}.txt', 'r') as chunk:
        for line in chunk:
            error = line.split()[-1].strip('""')
            error_dictionary[error] += 1


# 4. founds the N most common Error
def get_N_common_errors(errors, N):
    """
        :param errors: error's dictionary
        :param n: number of errors
        :return: N common errors appeared in the file
    """
    return errors.most_common(N)


def process_logs(file, N):
    """
       Processes the log file, counting occurrences of different error messages.

       :param file: Path to the log file.
       :param N: Number of most common errors to return.
       :return: List of (error, count) tuples.
       """
    if N <= 0:
        return "can't get negative number"

    errors_dict = Counter()
    num_of_files = split_lines(file)

    # counts the numbers of occurrence for every error in each file
    for i in range(num_of_files):
        count_occurrences_in_file(errors_dict, i)

    return get_N_common_errors(errors_dict, N)


# 5. space & time complexity:
# Definitions:
# * R = number of rows in the Excel file
# * C = number of rows per chunk (constant: LINES_PER_CHUNK)
# * F = number of chunk files (F ≈ R / C)
# * E = number of unique error messages
# * N = number of top errors requested

# - reading the file and writing into chunks files is O(R) - total O(R) Runtime
# - keeping in the memory one file at most O(C)
# - Storing E unique Error keys and their count - O(E) in Space complexity
# - get N common error (but internally sorting ) - O(E log E) Runtime

# so we'll get in Runtime: O(R + E logE)
# in Space: O(C + E)


"""
*****************************************************
section B: time_series
*****************************************************
"""


def clean_data(input_file, output_file):
    """
        Removes:
        - Duplicate rows.
        - Invalid 'timestamp' format (must start with DD/MM/YYYY).
        - Non-numeric 'value' column  (another check before processing).
    """
    seen = set()

    workbook = load_workbook(filename=input_file, read_only=True)
    sheet = workbook.active

    with open(output_file, 'w', newline='', encoding='utf-8') as outfile:
        writer = csv.writer(outfile)
        writer.writerow(['timestamp', 'value'])

        for row in sheet.iter_rows(min_row=2, values_only=True):
            timestamp = row[0]
            value = row[1]

            if timestamp is None or value is None:
                continue

            # Convert timestamp to string
            timestamp_str = timestamp.strftime('%d/%m/%Y %H:%M')

            # Validate date format
            if not is_valid_date(timestamp_str):
                continue

            # Check value is numeric and not "NaN"
            value_str = str(value).strip()
            if not is_number(value_str) or value_str == 'NaN':
                continue

            row_tuple = (timestamp_str, value_str)
            if row_tuple not in seen:
                seen.add(row_tuple)
                writer.writerow([timestamp_str, value_str])


def is_number(s):
    """
        checks if a given value is a number.
        :param s:
        :return: true if it's a number - false otherwise.
    """
    try:
        float(s)
        return True
    except ValueError:
        return False


def is_valid_date(timestamp_str):
    """
        checks if a given string is in a valid format of DD/MM/Y.
        :param timestamp_str:
        :return: True - if it's valid string, False - otherwise.
    """
    try:
        # Split out the date part before the space
        date_part = timestamp_str.strip().split()[0]
        datetime.strptime(date_part, "%d/%m/%Y")
        return True
    except (ValueError, IndexError):
        return False


def calculate_average(file):
    """
        Reads a CSV file containing timestamps and values, calculates the average
        value for each day and full hour.

        Parameter:
        - file: Path to the input CSV file.
    """
    daily_hour_values = defaultdict(list)  # {(date, time) : [value]}
    # making a dictionary for daily hours
    with open(file, 'r', newline='') as infile:
        reader = csv.DictReader(infile)

        for row in reader:
            # Parse the timestamp, assuming format is "day/month/year hour:minute:second"
            timestamp = datetime.strptime(row['timestamp'], '%d/%m/%Y %H:%M')

            # Round down to the full hour
            date_hour_key = (timestamp.date(), timestamp.hour)
            daily_hour_values[date_hour_key].append(float(row['value']))

    # calculate the averages
    averages = []
    for (date, hour), value in daily_hour_values.items():
        average = sum(value) / len(value)
        averages.append((date, hour, average))
    averages.sort()  # by day and time
    return averages


def split_by_day(file):
    """
    Splits a CSV time series file into separate files by day.

    Parameters:
    - file (str): Path to the cleaned CSV file.

    Output:
    - One CSV file per day, named YYYY-MM-DD.csv
    """
    file_handlers = {}
    writers_to_file = {}

    with open(file, 'r', newline='', encoding='utf-8') as infile:
        reader = csv.DictReader(infile)

        for row in reader:
            timestamp = datetime.strptime(row['timestamp'], '%d/%m/%Y %H:%M')
            date_str = timestamp.strftime('%Y-%m-%d')
            file_name = f"{date_str}.csv"

            if date_str not in writers_to_file:
                f = open(file_name, 'w', newline='', encoding='utf-8')
                file_handlers[date_str] = f
                writers_to_file[date_str] = csv.writer(f)
                writers_to_file[date_str].writerow(['timestamp', 'value'])  # Write header

            writers_to_file[date_str].writerow([row['timestamp'], row['value']])

    # Close all open files
    for f in file_handlers.values():
        f.close()


def process_all_daily_files(final_output_file):
    """
    Processes all daily files (e.g., 2025-06-28.csv), calculates hourly averages,
    and writes all results to a combined output file.

    Parameters:
    - final_output_file (str): Path to save the final averaged CSV file.
    """
    all_averages = []

    for file in os.listdir():
        if file.endswith('.csv') and file[:4].isdigit():  # e.g. 2025-06-28.csv
            averages = calculate_average(file)
            all_averages.extend(averages)

    all_averages.sort()  # sort all together

    with open(final_output_file, 'w', newline='', encoding='utf-8') as avgFile:
        writer = csv.writer(avgFile)
        writer.writerow(['Timestamp', 'Average'])

        for date, hour, avg in all_averages:
            timestamp_str = f"{date.strftime('%d/%m/%Y')} {hour:02d}:00"
            writer.writerow([timestamp_str, avg])


##########################################

stream_hourly_avg = defaultdict(lambda: {'sum': 0.0, 'count': 0})


def streaming_values(date_str, value):
    """
        Updates live hourly averages from streamed timestamp-value pairs.
        Keeps running sum and count for each (day, hour) key.

        Parameters:
        - date_str (str): Timestamp string in DD/MM/YYYY HH:MM format.
        - value (str): Numeric value as string.

        Returns:
        - Tuple[(date, hour), float]: The hour key and current average.
        """
    timestamp = datetime.strptime(date_str, "%d/%m/%Y %H:%M")
    key = (timestamp.date(), datetime.hour)
    value = float(value)

    # in this way we get the average on live
    stream_hourly_avg[key]['sum'] = + value
    stream_hourly_avg[key]['count'] += 1

    avg = stream_hourly_avg[key]['sum'] / stream_hourly_avg[key]['count']
    return key, avg


if __name__ == '__main__':
    # ********************************
    # section A : running code
    # ********************************
    # print(process_logs("logs.txt.xlsx", 5))

    # ********************************
    # section B : running code
    # ********************************
    # 1. preprocess the data before process it
    clean_data("time_series.xlsx", "clean_data.csv")  # takes almost one minute to be finished
    print("finish clean data")

    # # 2. calculating the average per daily hour
    # print(calculate_average("clean_data.csv"))
    # print("finish calculating")

    # 3. splitting time_series.csv into smaller parts, calculate the avg for each file and combine all
    # results together into one .csv file
    split_by_day("clean_data.csv")
    print("finish splitting")
    process_all_daily_files("combined_avg.csv")
    print("finish processing")

    # 4. if the data is giving in a streaming way instead from file, we will process the averages on live
    # as follow:
    # Instead of storing all the values for each hour, we only keep:
    #    - The sum of the values for each hour.
    #    - The count (number of values) that arrived for that hour
    # define default dict globally (stream_hourly_avg).
    # function streaming_values gets a date and a value (as strings) and update the hourly averages on
    # time.
    # by calling to streaming_values with different values and than printing stream_hourly_avg dictionary
    # we can see the averages gets updates on live.
    # streaming_values("03/10/2024 20:15", "15.3") - for example

    # 5. adjusting the code to support .parquet format:
    # the given file - time_series.parquet containing (timestamp, means_value, median_value, std_dev,
    # counts) columns rather than (timestamp, values) columns that is mentions in the exercise file.
    # i don't know if that a mistake or not, if so, inorder to adjust the code i would
    # have preform this code:
    # - cleaning the data - dropping rows with missing values/duplicates rows/non-numeric value
    # - group the row by hours (for example, turning timestamp to become the index)
    # - calculate the mean by : df.resample('H').mean()
    # - output thr results to csv file
    # otherwise, all is needed it to extract the timestamp and the mean_value columns to a csv file.

    # parquet is a binary, columnar storage format
    # the benefits of storing the data in parquet format over csv are:
    # 1. fast data access - only necessary column are read from the disk and not the whole file.
    # 2. the files are smaller compared to csv files, which reduces disk usage.
    # 3. Handling missing values - missing values are natively supported and efficiently stored, compared
    # to csv, where missing values needs to be checks and.
    # 4. parquet files store metadata such as column names and data type - this makes it easier to
    # load and understand data across different system.
